from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook


def build_sample_dpc_workbook_bytes(
    *,
    include_icd_sheet: bool = True,
    include_point_sheet: bool = True,
    include_point_label_column: bool = True,
    first_point_dpc_code: str = "010010xx9900xx",
) -> bytes:
    """Build a minimal official DPC workbook fixture as xlsx bytes."""

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    if include_icd_sheet:
        icd_sheet = workbook.create_sheet("４）ＩＣＤ")
        icd_sheet["A1"] = "MDCｺｰﾄﾞ"
        icd_sheet["B1"] = "分類ｺｰﾄﾞブンルイ"
        icd_sheet["D1"] = "ICDｺｰﾄﾞ"
        icd_sheet["A3"] = "01"
        icd_sheet["B3"] = "0010"
        icd_sheet["D3"] = "C700"

    if include_point_sheet:
        point_sheet = workbook.create_sheet("11）診断群分類点数表")
        point_sheet["B3"] = "番号バンゴウ"
        point_sheet["C3"] = "診断群分類番号"
        if include_point_label_column:
            point_sheet["D3"] = "傷病名ショウビョウメイ"
        point_sheet["B5"] = "1"
        point_sheet["C5"] = first_point_dpc_code
        if include_point_label_column:
            point_sheet["D5"] = "脳腫瘍"
        point_sheet["B6"] = "2"
        point_sheet["C6"] = "010010xx9901xx"
        if include_point_label_column:
            point_sheet["D6"] = "脳腫瘍"

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()

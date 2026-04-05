from __future__ import annotations

import csv
import json
import os
import re
from collections.abc import Iterable, Sequence
from pathlib import Path
from typing import Protocol
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.workbook.workbook import Workbook

from marume_data.fetch import resolve_latest_asset_path


RULES_CSV_HEADERS = (
    "rule_id",
    "priority",
    "dpc_code",
    "mdc_code",
    "label",
    "main_diagnosis",
    "procedures",
)

ICD_SHEET_NAME = "４）ＩＣＤ"
POINT_TABLE_SHEET_NAME = "11）診断群分類点数表"
POINT_TABLE_DPC_CODE_INDEX = 2
POINT_TABLE_LABEL_INDEX = 3
ICD_TABLE_MDC_CODE_INDEX = 0
ICD_TABLE_CLASSIFICATION_CODE_INDEX = 1
ICD_TABLE_CODE_INDEX = 3
WORKBOOK_SUFFIXES = {".xlsx", ".xlsm"}
DPC_CODE_PATTERN = re.compile(r"^\d{6}")


class WorksheetLike(Protocol):
    """Minimal worksheet protocol used by workbook extraction helpers."""

    def iter_rows(self, *, min_row: int, values_only: bool) -> Iterable[Sequence[object]]: ...


def scaffold_rules_csv_from_manifest(manifest_path: Path, output_csv_path: Path) -> Path:
    """Extract rules CSV rows from the official DPC workbook referenced by a manifest."""

    source_path = resolve_latest_asset_path(manifest_path, kind="official")
    if source_path is None:
        raise ValueError("official DPC workbook was not found in manifest")
    if source_path.suffix.lower() not in WORKBOOK_SUFFIXES:
        raise ValueError(f"official DPC workbook is required, but got: {source_path.name}")
    return scaffold_rules_csv_from_workbook(workbook_path=source_path, output_csv_path=output_csv_path)


def scaffold_rules_csv_from_workbook(workbook_path: Path, output_csv_path: Path) -> Path:
    """Extract flattened rule rows from an official DPC workbook."""

    if not workbook_path.exists():
        raise FileNotFoundError(workbook_path)
    if workbook_path.suffix.lower() not in WORKBOOK_SUFFIXES:
        raise ValueError(f"workbook file must end with .xlsx or .xlsm: {workbook_path.name}")

    try:
        rows = _extract_flat_rule_rows(workbook_path)
    except (InvalidFileException, BadZipFile, OSError) as exc:
        raise ValueError(f"failed to parse workbook file: {workbook_path.name}") from exc
    output_csv_path.parent.mkdir(parents=True, exist_ok=True)
    tmp_output_csv_path = output_csv_path.with_name(f".{output_csv_path.name}.tmp")
    with tmp_output_csv_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(RULES_CSV_HEADERS)
        writer.writerows(rows)
        handle.flush()
        os.fsync(handle.fileno())
    tmp_output_csv_path.replace(output_csv_path)

    metadata_path = output_csv_path.with_suffix(".source.json")
    metadata = {
        "source_workbook": str(workbook_path),
        "output_csv": str(output_csv_path),
        "status": "extracted",
        "row_count": len(rows),
        "note": "Rows were extracted from the official DPC workbook. Procedure mapping is still simplified.",
    }
    tmp_metadata_path = metadata_path.with_name(f".{metadata_path.name}.tmp")
    with tmp_metadata_path.open("w", encoding="utf-8") as handle:
        handle.write(json.dumps(metadata, ensure_ascii=False, indent=2) + "\n")
        handle.flush()
        os.fsync(handle.fileno())
    tmp_metadata_path.replace(metadata_path)
    return output_csv_path


def _extract_flat_rule_rows(source_path: Path) -> list[tuple[str, int, str, str, str, str, str]]:
    """Extract a minimal flattened rule CSV from the official DPC workbook."""

    workbook = load_workbook(source_path, read_only=True, data_only=True)
    try:
        icd_by_classification = _load_first_icd_by_classification(_require_sheet(workbook, ICD_SHEET_NAME))
        score_sheet = _require_sheet(workbook, POINT_TABLE_SHEET_NAME)
        rows: list[tuple[str, int, str, str, str, str, str]] = []
        priority = 10
        for row_number, row in enumerate(score_sheet.iter_rows(min_row=5, values_only=True), start=1):
            sheet_row_index = row_number + 4
            dpc_code = _normalize_cell(
                _cell_at(
                    row,
                    POINT_TABLE_DPC_CODE_INDEX,
                    sheet_name=POINT_TABLE_SHEET_NAME,
                    row_index=sheet_row_index,
                )
            )
            if not dpc_code:
                continue
            dpc_code_parts = _parse_dpc_code(dpc_code)
            if dpc_code_parts is None:
                continue
            label = _normalize_cell(
                _cell_at(
                    row,
                    POINT_TABLE_LABEL_INDEX,
                    sheet_name=POINT_TABLE_SHEET_NAME,
                    row_index=sheet_row_index,
                )
            ) or dpc_code
            mdc_code, classification_code = dpc_code_parts
            main_diagnosis = icd_by_classification.get((mdc_code, classification_code), "")
            rule_number = len(rows) + 1
            rows.append(
                (
                    f"R-{mdc_code}{classification_code}-{rule_number:05d}",
                    priority,
                    dpc_code,
                    mdc_code,
                    label,
                    main_diagnosis,
                    "",
                )
            )
            priority += 10
        return rows
    finally:
        workbook.close()


def _load_first_icd_by_classification(sheet: WorksheetLike) -> dict[tuple[str, str], str]:
    """Load the first ICD code for each classification from the ICD sheet."""

    mapping: dict[tuple[str, str], str] = {}
    for row_index, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
        mdc_code = _normalize_cell(
            _cell_at(row, ICD_TABLE_MDC_CODE_INDEX, sheet_name=ICD_SHEET_NAME, row_index=row_index)
        )
        classification_code = _normalize_cell(
            _cell_at(
                row,
                ICD_TABLE_CLASSIFICATION_CODE_INDEX,
                sheet_name=ICD_SHEET_NAME,
                row_index=row_index,
            )
        )
        icd_code = _normalize_cell(
            _cell_at(row, ICD_TABLE_CODE_INDEX, sheet_name=ICD_SHEET_NAME, row_index=row_index)
        )
        if not mdc_code or not classification_code or not icd_code:
            continue
        mdc_code = mdc_code.zfill(2)
        classification_code = classification_code.zfill(4)
        mapping.setdefault((mdc_code, classification_code), icd_code)
    return mapping


def _require_sheet(workbook: Workbook, sheet_name: str) -> WorksheetLike:
    """Return a worksheet by name or raise a clear error."""

    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"required sheet was not found in workbook: {sheet_name}")
    return workbook[sheet_name]


def _cell_at(row: Sequence[object], index: int, *, sheet_name: str, row_index: int) -> object:
    """Return one cell value or raise a clear error for malformed workbook rows."""

    if len(row) <= index:
        raise ValueError(
            f"malformed row in sheet {sheet_name}: row {row_index} does not have column index {index}"
        )
    return row[index]


def _normalize_cell(value: object) -> str:
    """Normalize workbook cell values into stripped strings."""

    if value is None:
        return ""
    return str(value).strip()


def _parse_dpc_code(dpc_code: str) -> tuple[str, str] | None:
    """Return MDC and classification codes when a DPC code starts with six digits."""

    if DPC_CODE_PATTERN.match(dpc_code) is None:
        return None
    return dpc_code[:2], dpc_code[2:6]
